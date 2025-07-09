import pandas as pd
from typing import List, Dict, Any, Optional
from app.schemas.test_import import TestImportData, QuestionImportData
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelTestParser:
    """Парсер для Excel файлов с тестами"""
    
    def __init__(self):
        self.required_columns = [
            'question_text', 'question_abstract', 'question_type', 
            'options', 'correct_answers', 'position'
        ]
    
    def parse_excel_file(self, file_path: str) -> TestImportData:
        """
        Парсит Excel файл и возвращает данные для импорта теста
        
        Args:
            file_path: Путь к Excel файлу
            
        Returns:
            TestImportData: Данные для импорта теста
            
        Raises:
            ValueError: При ошибке парсинга файла
        """
        try:
            logger.info(f"Начинаем парсинг Excel файла: {file_path}")
            
            # Используем контекстный менеджер для гарантированного закрытия файла
            with pd.ExcelFile(file_path) as excel_file:
                # Читаем метаданные теста (первый лист)
                test_metadata = pd.read_excel(excel_file, sheet_name=0)
                metadata = self._parse_test_metadata(test_metadata)
                
                # Ищем лист с вопросами
                questions_sheet = None
                for sheet_name in excel_file.sheet_names[1:]:  # Пропускаем первый лист
                    try:
                        questions_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        if 'question_text' in questions_df.columns:
                            questions_sheet = questions_df
                            logger.info(f"Найден лист с вопросами: {sheet_name}")
                            break
                    except Exception as e:
                        logger.warning(f"Не удалось прочитать лист {sheet_name}: {e}")
                        continue
                
                if questions_sheet is None:
                    # Если не нашли лист с вопросами, пробуем второй лист
                    if len(excel_file.sheet_names) > 1:
                        questions_sheet = pd.read_excel(excel_file, sheet_name=1)
                        logger.info("Используем второй лист как лист с вопросами")
                    else:
                        raise ValueError("Не найден лист с вопросами")
            
            # Файл уже закрыт, теперь парсим вопросы
            questions = self._parse_questions(questions_sheet)
            
            # Создаем TestImportData с вопросами
            metadata['questions'] = questions
            logger.info(f"Успешно распарсено {len(questions)} вопросов")
            return TestImportData(**metadata)
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге Excel файла: {str(e)}")
            raise ValueError(f"Ошибка при парсинге Excel файла: {str(e)}")
    
    def _parse_test_metadata(self, df: pd.DataFrame) -> dict:
        """
        Парсит метаданные теста из первого листа
        
        Args:
            df: DataFrame с метаданными
            
        Returns:
            dict: Словарь с метаданными теста
            
        Raises:
            ValueError: При отсутствии обязательных полей
        """
        metadata = {}
        # Соответствие русских и английских названий
        field_map = {
            'title': ['title', 'название теста'],
            'description': ['description', 'описание теста'],
            'category_name': ['category_name', 'категория'],
            'locale_code': ['locale_code', 'язык', 'локаль'],
            'time_limit': ['time_limit', 'лимит времени (сек)', 'лимит времени', 'лимит времени (мин)'],
            'welcome_message': ['welcome_message', 'приветственное сообщение'],
            'goodbye_message': ['goodbye_message', 'прощальное сообщение'],
            'answer_view_mode': ['answer_view_mode', 'режим показа ответов']
        }
        
        # Маппинг значений для answer_view_mode
        answer_view_mode_map = {
            'только ответы пользователя': 'user_only',
            'ответы пользователя': 'user_only',
            'user_only': 'user_only',
            'не показывать': 'none',
            'не показывать ответы': 'none',
            'none': 'none',
            'ответы пользователя с правильными': 'user_and_correct',
            'пользователя с правильными': 'user_and_correct',
            'user_and_correct': 'user_and_correct'
        }
        
        for _, row in df.iterrows():
            key = str(row.iloc[0]).strip().lower()
            value = row.iloc[1] if len(row) > 1 else None
            if pd.isna(value):
                continue
                
            for meta_key, variants in field_map.items():
                if key in variants:
                    # Особая обработка для time_limit (мин → сек)
                    if meta_key == 'time_limit':
                        try:
                            val = int(value)
                            # Если указано в минутах, переводим в сек
                            if 'мин' in key:
                                val = val * 60
                            metadata[meta_key] = val
                        except (ValueError, TypeError):
                            logger.warning(f"Некорректное значение time_limit: {value}")
                            metadata[meta_key] = None
                    # Особая обработка для answer_view_mode
                    elif meta_key == 'answer_view_mode':
                        value_str = str(value).strip().lower()
                        if value_str in answer_view_mode_map:
                            metadata[meta_key] = answer_view_mode_map[value_str]
                        else:
                            logger.warning(f"Некорректное значение answer_view_mode: {value}. Используется значение по умолчанию 'user_only'")
                            metadata[meta_key] = 'user_only'
                    else:
                        metadata[meta_key] = str(value)
        
        # Проверяем обязательные поля
        if 'title' not in metadata:
            raise ValueError("Отсутствует обязательное поле 'title'")
        if 'category_name' not in metadata:
            raise ValueError("Отсутствует обязательное поле 'category_name'")
        
        # Устанавливаем значение по умолчанию для answer_view_mode, если не указано
        if 'answer_view_mode' not in metadata:
            metadata['answer_view_mode'] = 'user_only'
            
        logger.info(f"Метаданные теста: {metadata}")
        return metadata
    
    def _parse_questions(self, df: pd.DataFrame) -> List[QuestionImportData]:
        """
        Парсит вопросы из второго листа
        
        Args:
            df: DataFrame с вопросами
            
        Returns:
            List[QuestionImportData]: Список вопросов
            
        Raises:
            ValueError: При ошибке парсинга вопросов
        """
        questions = []
        # Соответствие русских и английских названий
        col_map = {
            'question_text': ['question_text', 'текст вопроса'],
            'question_abstract': ['question_abstract', 'краткий текст'],
            'question_type': ['question_type', 'тип вопроса'],
            'options': ['options', 'варианты ответов (через |)', 'варианты ответов'],
            'correct_answers': ['correct_answers', 'правильный ответ', 'правильные ответы'],
            'max_score': ['max_score', 'максимальный балл'],
            'time_limit': ['time_limit', 'лимит времени (сек)', 'лимит времени']
        }
        # Русско-английское сопоставление типа вопроса
        type_map = {
            'одиночный выбор': 'SINGLE_CHOICE',
            'множественный выбор': 'MULTIPLE_CHOICE',
            'single_choice': 'SINGLE_CHOICE',
            'multiple_choice': 'MULTIPLE_CHOICE',
            'single choice': 'SINGLE_CHOICE',
            'multiple choice': 'MULTIPLE_CHOICE',
        }
        
        # Построим отображение: ключ -> колонка в df
        col_idx = {}
        for key, variants in col_map.items():
            for v in variants:
                for col in df.columns:
                    if str(col).strip().lower() == v:
                        col_idx[key] = col
                        break
                if key in col_idx:
                    break
        
        logger.info(f"Найдены колонки: {col_idx}")
        
        for idx, row in df.iterrows():
            try:
                # Пропускаем пустые строки
                qtext = row.get(col_idx.get('question_text', ''), '')
                if pd.isna(qtext) or str(qtext).strip() == '':
                    continue
                    
                # Тип вопроса: поддержка русских и англ. значений
                raw_type = str(row.get(col_idx.get('question_type', ''), '')).strip().lower()
                question_type = type_map.get(raw_type, raw_type.upper())
                
                question_data = {
                    'question_text': f'<p>{str(qtext).strip()}</p>',
                    'question_abstract': str(qtext).strip(),
                    'question_type': question_type,
                    'position': idx + 1,  # автоинкремент
                    'required': True,     # всегда True
                    'max_score': int(row.get(col_idx.get('max_score', ''), 1)),
                    'time_limit': int(row.get(col_idx.get('time_limit', ''), 0)) if pd.notna(row.get(col_idx.get('time_limit', ''))) else 0
                }
                
                # Парсим варианты ответов
                options_str = str(row.get(col_idx.get('options', ''), ''))
                options = self._parse_options(options_str)
                question_data['options'] = options
                
                # Парсим правильные ответы
                correct_answers_str = str(row.get(col_idx.get('correct_answers', ''), ''))
                correct_answers = self._parse_correct_answers(correct_answers_str, options)
                question_data['correct_answers'] = correct_answers
                
                questions.append(QuestionImportData(**question_data))
                
            except Exception as e:
                logger.error(f"Ошибка при парсинге вопроса в строке {idx + 2}: {str(e)}")
                raise ValueError(f"Ошибка при парсинге вопроса в строке {idx + 2}: {str(e)}")
                
        return questions
    
    def _parse_options(self, options_str: str) -> List[str]:
        """
        Парсит варианты ответов из строки
        
        Args:
            options_str: Строка с вариантами ответов
            
        Returns:
            List[str]: Список вариантов ответов
        """
        if not options_str or options_str.strip() == '':
            return []
        
        # Разделяем по символу '|' или переносу строки
        options = []
        for option in re.split(r'[|\n]', options_str):
            option = option.strip()
            if option:
                # Убираем маркеры типа ~ или =
                option = re.sub(r'^[~=]\s*', '', option)
                options.append(option)
        
        return options
    
    def _parse_correct_answers(self, correct_answers_str: str, options: List[str]) -> List[str]:
        """
        Парсит правильные ответы
        
        Args:
            correct_answers_str: Строка с правильными ответами
            options: Список вариантов ответов
            
        Returns:
            List[str]: Список правильных ответов
        """
        if not correct_answers_str or correct_answers_str.strip() == '':
            return []
        
        correct_answers = []
        
        # Сначала проверяем, является ли это индексом (число)
        if correct_answers_str.strip().isdigit():
            idx = int(correct_answers_str.strip()) - 1  # Excel индексы начинаются с 1
            if 0 <= idx < len(options):
                correct_answers.append(options[idx])
        else:
            # Если это текст ответа, ищем точное совпадение среди вариантов
            answer_text = correct_answers_str.strip()
            
            # Ищем точное совпадение среди вариантов
            found = False
            for option in options:
                if option.strip() == answer_text:
                    correct_answers.append(option)
                    found = True
                    break
            
            # Если не нашли точное совпадение, пробуем разделить по |
            if not found and '|' in correct_answers_str:
                for answer_part in correct_answers_str.split('|'):
                    answer_part = answer_part.strip()
                    if answer_part:
                        for option in options:
                            if option.strip() == answer_part:
                                correct_answers.append(option)
                                break
        
        return correct_answers


def create_excel_template(file_path_or_buffer) -> None:
    """
    Создает базовый шаблон Excel файла для импорта тестов
    
    Args:
        file_path_or_buffer: Путь к файлу или BytesIO объект
    """
    try:
        logger.info(f"Создание Excel шаблона")
        
        # Создаем Excel файл с двумя листами
        with pd.ExcelWriter(file_path_or_buffer, engine='openpyxl') as writer:
            # Лист 1: Метаданные теста
            metadata_df = pd.DataFrame([
                ['title', 'Название теста'],
                ['description', 'Описание теста'],
                ['category_name', 'Название категории'],
                ['locale_code', 'ru'],
                ['time_limit', 1800],
                ['welcome_message', 'Добро пожаловать в тест!'],
                ['goodbye_message', 'Спасибо за прохождение теста!'],
                ['answer_view_mode', 'только ответы пользователя']
            ], columns=['Field', 'Value'])
            metadata_df.to_excel(writer, sheet_name='Метаданные теста', index=False)
            
            # Лист 2: Вопросы
            questions_df = pd.DataFrame([
                ['Текст вопроса?', 'Краткий текст', 'SINGLE_CHOICE', 'Вариант 1|Вариант 2|Вариант 3', 'Вариант 1', 1, 0, 1, True],
                ['Другой вопрос?', 'Краткий текст 2', 'MULTIPLE_CHOICE', 'Вариант A|Вариант B|Вариант C', 'Вариант A|Вариант B', 2, 0, 2, True]
            ], columns=['question_text', 'question_abstract', 'question_type', 'options', 'correct_answers', 'max_score', 'time_limit', 'position', 'required'])
            questions_df.to_excel(writer, sheet_name='Вопросы', index=False)
            
        logger.info("Excel шаблон успешно создан")
        
    except Exception as e:
        logger.error(f"Ошибка при создании Excel шаблона: {str(e)}")
        raise ValueError(f"Ошибка при создании Excel шаблона: {str(e)}")


def create_friendly_excel_template(file_path_or_buffer) -> None:
    """
    Создает дружелюбный шаблон Excel с русскими названиями
    
    Args:
        file_path_or_buffer: Путь к файлу или BytesIO объект
    """
    try:
        logger.info(f"Создание дружелюбного Excel шаблона")
        
        wb = openpyxl.Workbook()
        
        # Лист 1: Метаданные теста
        ws1 = wb.active
        ws1.title = 'Метаданные теста'
        
        metadata = [
            ['Название теста', 'Введите название теста'],
            ['Описание теста', 'Введите описание теста'],
            ['Категория', 'Введите название категории'],
            ['Язык', 'ru'],
            ['Лимит времени (сек)', '1800'],
            ['Приветственное сообщение', 'Добро пожаловать в тест!'],
            ['Прощальное сообщение', 'Спасибо за прохождение теста!'],
            ['Режим показа ответов', 'только ответы пользователя']
        ]
        
        for row in metadata:
            ws1.append(row)
        
        # Выделяем заголовки
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        
        # Добавляем валидацию для режима показа ответов
        dv_answer_mode = DataValidation(type="list", formula1='"только ответы пользователя,не показывать,ответы пользователя с правильными"', allow_blank=True)
        dv_answer_mode.add('H2:H100')
        ws1.add_data_validation(dv_answer_mode)
        
        # Лист 2: Вопросы
        ws2 = wb.create_sheet('Вопросы')
        columns = [
            'Текст вопроса', 'Тип вопроса', 'Варианты ответов (через |)',
            'Правильный ответ', 'Максимальный балл', 'Лимит времени (сек)'
        ]
        ws2.append(columns)
        
        # Выделяем заголовки
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        
        # Добавляем примеры вопросов
        examples = [
            ['Как компания «Сэльвин» расценивает стиль одежды для сотрудников?', 'Одиночный выбор', 'Строго официальный|Свободный и удобный|Спортивный|В зависимости от сезона', 'Свободный и удобный', 1, 0],
            ['Какие качества важны для сотрудника?', 'Множественный выбор', 'Ответственность|Креативность|Пунктуальность|Коммуникабельность', 'Ответственность|Пунктуальность', 2, 0]
        ]
        
        for example in examples:
            ws2.append(example)
        
        # Добавляем валидацию для типа вопроса
        dv = DataValidation(type="list", formula1='"Одиночный выбор,Множественный выбор"', allow_blank=True)
        dv.add(f'C2:C{len(examples) + 1}')
        ws2.add_data_validation(dv)
        
        # Сохраняем файл
        wb.save(file_path_or_buffer)
        logger.info("Дружелюбный Excel шаблон успешно создан")
        
    except Exception as e:
        logger.error(f"Ошибка при создании дружелюбного Excel шаблона: {str(e)}")
        raise ValueError(f"Ошибка при создании дружелюбного Excel шаблона: {str(e)}")


def create_full_friendly_excel_template(file_path_or_buffer) -> None:
    """
    Создает полный дружелюбный шаблон Excel с тремя листами
    
    Args:
        file_path_or_buffer: Путь к файлу или BytesIO объект
    """
    try:
        logger.info(f"Создание полного дружелюбного Excel шаблона")
        
        wb = openpyxl.Workbook()
        
        # 1. Основные настройки
        ws_settings = wb.active
        ws_settings.title = 'Основные настройки'
        settings = [
            ('Название теста', 'Введите название теста'),
            ('Описание теста', 'Введите описание теста'),
            ('Категория', 'Введите название категории'),
            ('Язык', 'ru'),
            ('Лимит времени (сек)', '1800'),
            ('Приветственное сообщение', 'Добро пожаловать в тест!'),
            ('Прощальное сообщение', 'Спасибо за прохождение теста!'),
            ('Режим показа ответов', 'только ответы пользователя')
        ]
        ws_settings.append(['Поле', 'Значение'])
        for row in settings:
            ws_settings.append(row)
        
        # Выделяем заголовки
        for cell in ws_settings[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # Добавляем валидацию для режима показа ответов
        dv_answer_mode = DataValidation(type="list", formula1='"только ответы пользователя,не показывать,ответы пользователя с правильными"', allow_blank=True)
        dv_answer_mode.add('H3:H100')
        ws_settings.add_data_validation(dv_answer_mode)
        
        # 2. Вопросы
        ws = wb.create_sheet('Вопросы')
        columns = [
            'Текст вопроса', 'Тип вопроса', 'Варианты ответов (через |)',
            'Правильный ответ', 'Максимальный балл', 'Лимит времени (сек)'
        ]
        ws.append(columns)
        
        # Выделяем заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # Добавляем примеры
        examples = [
            ['Как компания «Сэльвин» расценивает стиль одежды для сотрудников?', 'Одиночный выбор', 'Строго официальный|Свободный и удобный|Спортивный|В зависимости от сезона', 'Свободный и удобный', 1, 0],
            ['Какие качества важны для сотрудника?', 'Множественный выбор', 'Ответственность|Креативность|Пунктуальность|Коммуникабельность', 'Ответственность|Пунктуальность', 2, 0]
        ]
        
        for example in examples:
            ws.append(example)
        
        # Добавляем валидацию для типа вопроса
        dv = DataValidation(type="list", formula1='"Одиночный выбор,Множественный выбор"', allow_blank=True)
        dv.add(f'C2:C{len(examples) + 1}')
        ws.add_data_validation(dv)
        
        # 3. Инструкции
        ws_instructions = wb.create_sheet('Инструкции')
        
        instructions = [
            'Как заполнять шаблон для импорта тестов:',
            '1. Заполните лист "Основные настройки" — это информация о тесте.',
            '2. На листе "Вопросы" заполните каждый вопрос в отдельной строке.',
            '3. Обязательные поля выделены цветом.',
            '4. Варианты ответов указывайте через | (вертикальная черта).',
            '5. Для одиночного выбора правильный ответ — один, для множественного — через |.',
            '6. Тип вопроса выбирайте из выпадающего списка.',
            '7. Режим показа ответов выбирайте из выпадающего списка:',
            '   • "только ответы пользователя" — показывать только ответы пользователя',
            '   • "не показывать" — не показывать ответы вообще',
            '   • "ответы пользователя с правильными" — показывать ответы пользователя с правильными ответами',
            '',
            'Пример вопроса:',
            'Текст вопроса: Как компания «Сэльвин» расценивает стиль одежды для сотрудников?',
            'Варианты: Строго официальный|Свободный и удобный|Спортивный|В зависимости от сезона',
            'Правильный ответ: Свободный и удобный',
            'Тип вопроса: Одиночный выбор',
            'Максимальный балл: 1',
            'Лимит времени (сек): 0',
            '',
            'Пояснения по полям:',
            'Текст вопроса — формулировка для пользователя.',
            'Тип вопроса — одиночный или множественный выбор.',
            'Варианты ответов — список вариантов через |.',
            'Правильный ответ — правильный вариант или варианты через |.',
            'Максимальный балл — количество баллов за правильный ответ.',
            'Лимит времени (сек) — время на ответ в секундах (0 = без ограничений).',
            'Режим показа ответов — что показывать пользователю после прохождения теста.'
        ]
        
        for instruction in instructions:
            ws_instructions.append([instruction])
        
        # Выделяем заголовок инструкций
        ws_instructions['A1'].font = Font(bold=True, size=14)
        ws_instructions['A1'].fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        
        # Сохраняем файл
        wb.save(file_path_or_buffer)
        logger.info("Полный дружелюбный Excel шаблон успешно создан")
        
    except Exception as e:
        logger.error(f"Ошибка при создании полного дружелюбного Excel шаблона: {str(e)}")
        raise ValueError(f"Ошибка при создании полного дружелюбного Excel шаблона: {str(e)}")


def generate_analytics_excel(analytics_data: List[Dict[str, Any]]) -> BytesIO:
    """
    Формирует и стилизует Excel-отчёт по аналитическим данным (pivot, порядок, цвета, сумма баллов и т.д.).
    Возвращает BytesIO с готовым Excel-файлом.
    """
    df = pd.DataFrame(analytics_data)
    if 'user_full_name' in df.columns:
        # Обрабатываем случаи с None/NaN значениями
        df['user_full_name'] = df['user_full_name'].fillna('')
        df[['Имя', 'Фамилия']] = df['user_full_name'].str.split(' ', n=1, expand=True)
        # Заполняем NaN значения пустыми строками
        df['Имя'] = df['Имя'].fillna('')
        df['Фамилия'] = df['Фамилия'].fillna('')
    else:
        df['Имя'] = ''
        df['Фамилия'] = ''
    # Проверяем, что у нас есть необходимые колонки
    required_columns = ['test_title', 'user_test_score', 'test_max_score', 'test_percent']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Отсутствуют необходимые колонки: {missing_columns}")
    
    # Проверяем, что у нас есть данные
    if df.empty:
        raise ValueError("Нет данных для создания отчёта")
    
    pivot = df.pivot_table(
        index=['Имя', 'Фамилия'],
        columns='test_title',
        values=['user_test_score', 'test_max_score', 'test_percent'],
        aggfunc='first'
    )
    
    # Проверяем, что pivot table не пустой
    if pivot.empty:
        raise ValueError("Не удалось создать сводную таблицу. Возможно, нет данных для отображения.")
    
    pivot = pivot.reorder_levels([1, 0], axis=1).sort_index(axis=1, level=0)
    pivot = pivot.reset_index()
    pivot.columns = [
        f"{col[0]} ({col[1]})" if isinstance(col, tuple) and col[1] else str(col[0])
        for col in pivot.columns
    ]
    score_cols = [col for col in pivot.columns if col.endswith('(user_test_score)')]
    pivot['Сумма баллов (личная)'] = pivot[score_cols].fillna(0).sum(axis=1)
    fam_idx = list(pivot.columns).index('Фамилия')
    cols = list(pivot.columns)
    cols.insert(fam_idx + 1, cols.pop(cols.index('Сумма баллов (личная)')))
    pivot = pivot[cols]
    rename_map = {}
    for col in pivot.columns:
        if col.endswith('(user_test_score)'):
            test = col.split(' (')[0]
            rename_map[col] = f"{test} (баллы)"
        elif col.endswith('(test_max_score)'):
            test = col.split(' (')[0]
            rename_map[col] = f"Вопросы {test}"
        elif col.endswith('(test_percent)'):
            test = col.split(' (')[0]
            rename_map[col] = f"{test} (уровень %)"
    pivot = pivot.rename(columns=rename_map)
    main_cols = ['Имя', 'Фамилия', 'Сумма баллов (личная)']
    test_names = set()
    for col in pivot.columns:
        if col not in main_cols:
            if col.endswith('(баллы)'):
                test_names.add(col[:-8])
            elif col.startswith('Вопросы '):
                test_names.add(col[8:])
            elif col.endswith('(уровень %)'):
                test_names.add(col[:-12])
    test_names = sorted(test_names)
    ordered_cols = main_cols.copy()
    for test in test_names:
        for suffix in ['(баллы)', 'Вопросы', '(уровень %)']:
            if suffix == '(баллы)':
                col = f'{test} (баллы)'
            elif suffix == 'Вопросы':
                col = f'Вопросы {test}'
            else:
                col = f'{test} (уровень %)' 
            if col in pivot.columns:
                ordered_cols.append(col)
    pivot = pivot[ordered_cols]
    for col in pivot.select_dtypes(include=['float']).columns:
        pivot[col] = pivot[col].round(2)
    def convert_percent(x):
        try:
            if x is None or pd.isna(x):
                return x
            return round(float(x)/100, 4)
        except (ValueError, TypeError):
            return x
    
    percent_cols = [col for col in pivot.columns if col.endswith('(уровень %)')]
    for col in percent_cols:
        pivot[col] = pivot[col].apply(convert_percent)
    output = BytesIO()
    sheet_name = 'Pivot User-Tests'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pivot.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        green_cols = []
        for idx, cell in enumerate(ws[1]):
            col_name = pivot.columns[idx]
            if col_name in ['Имя', 'Фамилия', 'Сумма баллов (личная)'] or col_name.endswith('(баллы)'):
                cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                if col_name == 'Сумма баллов (личная)' or col_name.endswith('(баллы)'):
                    green_cols.append(idx)
            elif col_name.startswith('Вопросы ') or col_name.endswith('(уровень %)'):
                cell.fill = PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for idx in green_cols:
            for row in ws.iter_rows(min_row=2, min_col=idx+1, max_col=idx+1, max_row=ws.max_row):
                for cell in row:
                    cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        for idx, cell in enumerate(ws[1]):
            if '(уровень %)' in str(cell.value):
                col_letter = get_column_letter(idx + 1)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = '0.00%'
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
        ws.auto_filter.ref = ws.dimensions
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 18
    output.seek(0)
    return output 