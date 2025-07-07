import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import re

from app.models import Result, UserAnswer, User, Test, Question


def transliterate_cyrillic(text: str) -> str:
    """Транслитерация кириллических символов в латиницу"""
    cyrillic_to_latin = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
    }
    
    result = ''
    for char in text:
        result += cyrillic_to_latin.get(char, char)
    
    # Заменяем все не-ASCII символы на подчеркивание
    result = re.sub(r'[^a-zA-Z0-9_\-\s]', '_', result)
    # Заменяем множественные подчеркивания на одно
    result = re.sub(r'_+', '_', result)
    # Убираем пробелы и заменяем на подчеркивания
    result = re.sub(r'\s+', '_', result)
    # Убираем начальные и конечные подчеркивания
    result = result.strip('_')
    
    return result


def strip_html(text):
    return re.sub(r'<[^>]+>', '', text or '').strip()


def generate_result_excel(result: Result, answers: List[UserAnswer]) -> BytesIO:
    """Генерация Excel отчета по результату теста"""
    output = BytesIO()
    
    # Создаем DataFrame с основной информацией о результате
    result_data = {
        'Поле': [
            'ID результата',
            'Пользователь',
            'Email пользователя',
            'Тест',
            'Балл',
            'Дата прохождения',
            'Время начала',
            'Время окончания',
            'Общее время (сек)',
            'Статус прохождения'
        ],
        'Значение': [
            result.id,
            result.user.name or 'Не указано',
            result.user.email or 'Не указано',
            result.test.title or 'Не указано',
            result.score or 0,
            result.created_at.strftime('%d.%m.%Y %H:%M:%S') if result.created_at else 'Не указано',
            result.result.get('time_start', '') if result.result else '',
            result.result.get('time_end', '') if result.result else '',
            result.result.get('total_time_seconds', '') if result.result else '',
            'Пройден' if result.result and result.result.get('is_passed', False) else 'Не пройден'
        ]
    }
    
    result_df = pd.DataFrame(result_data)
    
    # Формируем ответы из result.result['checked_answers']
    checked_answers = result.result.get('checked_answers', []) if result.result else []
    answers_data = []
    for i, ans in enumerate(checked_answers, 1):
        answers_data.append({
            '№': i,
            'Вопрос': strip_html(ans.get('question_text')),
            'Тип вопроса': get_question_type_ru(ans.get('question_type')),
            'Ответ пользователя': strip_html(ans.get('check_details', {}).get('user_answer')),
            'Правильный ответ': strip_html(ans.get('check_details', {}).get('correct_answer')),
            'Правильно': 'Да' if ans.get('is_correct') else 'Нет',
            'Баллы': ans.get('score', 0),
            'Время ответа (сек)': ans.get('time_seconds', 0)
        })
    
    answers_df = pd.DataFrame(answers_data)
    
    # Создаем Excel файл
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Лист с основной информацией
        result_df.to_excel(writer, sheet_name='Информация о результате', index=False)
        
        # Лист с ответами
        answers_df.to_excel(writer, sheet_name='Ответы на вопросы', index=False)
        
        # Форматирование листа с информацией
        worksheet_info = writer.sheets['Информация о результате']
        _format_info_worksheet(worksheet_info)
        
        # Форматирование листа с ответами
        worksheet_answers = writer.sheets['Ответы на вопросы']
        _format_answers_worksheet(worksheet_answers)
    
    output.seek(0)
    return output


def _format_info_worksheet(worksheet):
    """Форматирование листа с информацией о результате"""
    # Стили
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматирование заголовков
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Форматирование данных
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Автоматическая ширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _format_answers_worksheet(worksheet):
    """Форматирование листа с ответами на вопросы"""
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматирование заголовков
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Форматирование данных
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Автоматическая ширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width 


QUESTION_TYPE_RU = {
    'SINGLE_CHOICE': 'Одиночный выбор',
    'MULTIPLE_CHOICE': 'Множественный выбор',
    'TEXT': 'Текстовый ответ',
    'NUMBER': 'Числовой ответ',
    'BOOLEAN': 'Да/Нет',
    # Добавьте другие типы, если есть
}

def get_question_type_ru(qtype: str) -> str:
    if not qtype:
        return 'Не указан'
    return QUESTION_TYPE_RU.get(qtype.upper(), qtype) 